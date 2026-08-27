import copy
import os
import re
from pathlib import Path
from datetime import date, timedelta
from uuid import uuid4

from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy.exc import IntegrityError
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from .auth import create_access_token, get_current_user, hash_password, verify_password
from .database import Base, SessionLocal, engine, get_db
from .models import AnnualLeave, ContractFile, Employee, EmployeeDocument, SickLeave, User
from .schemas import (
    AnnualLeaveCreateRequest,
    AnnualLeaveResponse,
    AnnualLeaveUpdateRequest,
    ChangePasswordRequest,
    ContractFileResponse,
    EmployeeCreateRequest,
    EmployeeDocumentResponse,
    EmployeeProfileResponse,
    EmployeeLeaveReportResponse,
    EmployeeLeaveSummaryResponse,
    EmployeeResponse,
    LoginRequest,
    SickLeaveCreateRequest,
    SickLeaveResponse,
    SickLeaveUpdateRequest,
    TokenResponse,
    UserResponse,
    StartupChecklistRequest,
    EmployeeUpdateRequest,
    UniformIssueSaveRequest,
    UniformIssueWorkbookResponse,
)


app = FastAPI(title="Employee Management App")
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "/app/uploads"))
PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR_ENV = os.getenv("TEMPLATE_DIR")
UNIFORM_ISSUE_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
UNIFORM_CARE_LETTER_MIME_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
ALLOWED_DOCUMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".jpg", ".jpeg", ".xls", ".xlsx"}
ALLOWED_CONTRACT_EXTENSIONS = ALLOWED_DOCUMENT_EXTENSIONS
ALLOWED_MEDICAL_CERT_EXTENSIONS = {".pdf", ".doc", ".docx", ".jpg", ".jpeg", ".png"}
UNIFORM_ISSUE_MAX_ROWS = 10
UNIFORM_ISSUE_DESCRIPTION_OPTIONS = [
    "NAME TAG",
    "FUEL PUMP TAG",
    "SHIRT",
    "TROUSERS",
    "JACKET- LIGHT WEIGHT",
    "JACKET- PADDED(extra item)",
    "CAP",
    "BEANIE",
    "SAFTY SHOES",
]
DOCUMENT_TYPE_LABELS = {
    "contract": "Contract",
    "passport_id_copy": "Passport / ID Copy",
    "permit_copy": "Permit Copy",
    "bank_acc": "Bank Acc",
    "written_warning": "Written Warning",
    "general": "General",
    "uniform_issue": "Uniform Issue",
    "uniform_care_letter": "Uniform Care Letter",
}
UNIFORM_ISSUE_ROW_MAP = {
    "NAME TAG": 7,
    "FUEL PUMP TAG": 8,
    "SHIRT": 9,
    "TROUSERS": 10,
    "JACKET- LIGHT WEIGHT": 11,
    "JACKET- PADDED(extra item)": 12,
    "CAP": 13,
    "BEANIE": 14,
    "SAFTY SHOES": 15,
}
UNIFORM_ISSUE_ROW_MAP_BY_ROW = {row: description for description, row in UNIFORM_ISSUE_ROW_MAP.items()}
UNIFORM_CARE_LETTER_FILES = {
    "Forecourt": "UNIFORM CARE LETTER-Forecourt.docx",
    "Cashier": "UNIFORM CARE LETTER-Cashiers-Bakers.docx",
    "Baker": "UNIFORM CARE LETTER-Cashiers-Bakers.docx",
    "Car Wash": "UNIFORM CARE LETTER-Car-Wash.docx",
}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def seed_admin_user(db: Session) -> None:
    existing_user = db.scalar(select(User).where(User.username == "admin"))
    if existing_user is None:
        db.add(User(username="admin", password_hash=hash_password("admin")))
        db.commit()


DEFAULT_STARTUP_CHECKLIST = {
    "Contract": {"done": False, "na": False, "date": None},
    "Contract Status": {"done": False, "na": False, "date": None, "status": "", "info": "", "on_system": False},
    "Permit Status": {"done": False, "na": False, "date": None},
    "Passport / ID Copy": {"done": False, "na": False, "date": None},
    "Permit Copy": {"done": False, "na": False, "date": None},
    "Name for file & Name Badge": {"done": False, "na": False, "date": None},
    "Uniform & Uniform Care Letter": {"done": False, "na": False, "date": None},
    "Start": {"done": False, "na": False, "date": None},
    "UIF": {"done": False, "na": False, "date": None},
    "MIBCO reg": {"done": False, "na": False, "date": None},
    "Open Weekly Deduction Sheet": {"done": False, "na": False, "date": None},
    "Disciplinary Spreadsheet & Permit Spreadsheet": {"done": False, "na": False, "date": None},
    "Phone No. onto Contacts List": {"done": False, "na": False, "date": None},
    "Staff Instructions / Training Pack / Job Description": {"done": False, "na": False, "date": None},
}


def build_startup_defaults() -> dict:
    return copy.deepcopy(DEFAULT_STARTUP_CHECKLIST)


def normalize_startup_data(data: dict | None) -> dict:
    normalized = build_startup_defaults()
    if not data:
        return normalized

    legacy_job_description = data.get("Job Description") if isinstance(data, dict) else None
    legacy_uniform_row = data.get("Uniform & Uniform Letter & Uniform Sizes Onto Chart & Employ No. & Badge No.") if isinstance(data, dict) else None
    if (
        isinstance(legacy_job_description, dict)
        and "Staff Instructions / Training Pack / Job Description" not in data
    ):
        data = dict(data)
        data["Staff Instructions / Training Pack / Job Description"] = legacy_job_description
    if (
        isinstance(legacy_uniform_row, dict)
        and "Uniform & Uniform Care Letter" not in data
    ):
        data = dict(data)
        data["Uniform & Uniform Care Letter"] = legacy_uniform_row

    for key, value in data.items():
        if key in normalized and isinstance(value, dict):
            normalized[key] = {**normalized[key], **value}
        else:
            normalized[key] = value
    return normalized


@app.on_event("startup")
def startup_event() -> None:
    Base.metadata.create_all(bind=engine)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with engine.begin() as connection:
        connection.execute(
            text(
                "ALTER TABLE employees ADD COLUMN IF NOT EXISTS department VARCHAR(50)"
            )
        )
        connection.execute(
            text(
                "ALTER TABLE employees ADD COLUMN IF NOT EXISTS archived BOOLEAN NOT NULL DEFAULT FALSE"
            )
        )
        connection.execute(
            text(
                "ALTER TABLE employees DROP COLUMN IF EXISTS hire_date"
            )
        )
        connection.execute(
            text(
                "CREATE UNIQUE INDEX IF NOT EXISTS "
                "uq_employees_passport_id ON employees (passport_id)"
            )
        )
    db = SessionLocal()
    try:
        seed_admin_user(db)
    finally:
        db.close()


@app.get("/api/health")
def health_check() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/api/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> TokenResponse:
    user = db.scalar(select(User).where(User.username == payload.username))
    if user is None or not verify_password(payload.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password",
        )

    return TokenResponse(access_token=create_access_token(user))


@app.get("/api/me", response_model=UserResponse)
def read_current_user(current_user: User = Depends(get_current_user)) -> UserResponse:
    return current_user


@app.post("/api/change-password")
def change_password(
    payload: ChangePasswordRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    if current_user.username != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin can change password")

    if not verify_password(payload.current_password, current_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect",
        )

    current_user.password_hash = hash_password(payload.new_password)
    db.add(current_user)
    db.commit()

    return {"message": "Password updated successfully"}


@app.get("/api/employees", response_model=list[EmployeeResponse])
def list_employees(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[EmployeeResponse]:
    employees = db.scalars(
        select(Employee)
        .where(Employee.archived.is_(False))
        .order_by(Employee.name.asc())
    ).all()
    return employees


@app.get("/api/archived-employees", response_model=list[EmployeeResponse])
def list_archived_employees(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[EmployeeResponse]:
    employees = db.scalars(
        select(Employee)
        .where(Employee.archived.is_(True))
        .order_by(Employee.name.asc())
    ).all()
    return employees


@app.post("/api/employees", response_model=EmployeeResponse, status_code=status.HTTP_201_CREATED)
def create_employee(
    payload: EmployeeCreateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeResponse:
    passport_id = payload.passport_id.strip()
    existing = db.scalar(select(Employee).where(Employee.passport_id == passport_id))
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Passport/ID must be unique")

    employee = Employee(
        name=payload.name.strip(),
        department=payload.department.strip(),
        passport_id=passport_id,
        startup_data=build_startup_defaults(),
    )
    db.add(employee)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Passport/ID must be unique") from None
    db.refresh(employee)
    return employee


@app.get("/api/employees/{employee_id}", response_model=EmployeeResponse)
def get_employee(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeResponse:
    employee = db.get(Employee, employee_id)
    if employee is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
    return employee


@app.put("/api/employees/{employee_id}", response_model=EmployeeResponse)
def update_employee(
    employee_id: int,
    payload: EmployeeUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeResponse:
    employee = db.get(Employee, employee_id)
    if employee is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")

    passport_id = payload.passport_id.strip()
    existing = db.scalar(
        select(Employee).where(Employee.passport_id == passport_id, Employee.id != employee_id)
    )
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Passport/ID must be unique")

    employee.name = payload.name.strip()
    employee.department = payload.department.strip()
    employee.passport_id = passport_id
    db.add(employee)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Passport/ID must be unique") from None
    db.refresh(employee)
    return employee


@app.post("/api/employees/{employee_id}/archive", response_model=EmployeeResponse)
def archive_employee(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeResponse:
    employee = _get_employee_or_404(db, employee_id)
    employee.archived = True
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee


@app.post("/api/employees/{employee_id}/restore", response_model=EmployeeResponse)
def restore_employee(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeResponse:
    employee = _get_employee_or_404(db, employee_id)
    employee.archived = False
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee


@app.get("/api/employees/{employee_id}/startup")
def get_employee_startup(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    employee = db.get(Employee, employee_id)
    if employee is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
    return {
        "employee_id": employee.id,
        "employee_name": employee.name,
        "department": employee.department,
        "archived": employee.archived,
        "startup_data": normalize_startup_data(employee.startup_data),
    }


@app.put("/api/employees/{employee_id}/startup")
def update_employee_startup(
    employee_id: int,
    payload: StartupChecklistRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    employee = db.get(Employee, employee_id)
    if employee is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")

    employee.startup_data = normalize_startup_data(payload.checklist)
    db.add(employee)
    db.commit()
    return {"message": "Startup checklist updated successfully"}


def _get_employee_or_404(db: Session, employee_id: int) -> Employee:
    employee = db.get(Employee, employee_id)
    if employee is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
    return employee


def _contract_file_path(stored_filename: str) -> Path:
    return UPLOAD_DIR / stored_filename


def _document_path(stored_filename: str) -> Path:
    return UPLOAD_DIR / stored_filename


def _normalize_document_type(document_type: str) -> str:
    if document_type not in DOCUMENT_TYPE_LABELS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid document type")
    return document_type


def _uniform_issue_template_path() -> Path:
    candidates = []

    if TEMPLATE_DIR_ENV:
        candidates.append(Path(TEMPLATE_DIR_ENV) / "ISSUE-Uniforms.xlsx")

    candidates.extend(
        [
            PROJECT_ROOT / "Document Templates" / "ISSUE-Uniforms.xlsx",
            Path(__file__).resolve().parents[1] / "Document Templates" / "ISSUE-Uniforms.xlsx",
            Path("/app/Document Templates/ISSUE-Uniforms.xlsx"),
        ]
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return candidates[0]


def _uniform_care_letter_template_path(department: str | None) -> Path:
    if not department:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Employee department is missing")

    template_name = UNIFORM_CARE_LETTER_FILES.get(department.strip())
    if template_name is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported department")

    candidates = []
    if TEMPLATE_DIR_ENV:
        candidates.append(Path(TEMPLATE_DIR_ENV) / template_name)
    candidates.extend(
        [
            PROJECT_ROOT / "Document Templates" / template_name,
            Path(__file__).resolve().parents[1] / "Document Templates" / template_name,
            Path("/app/Document Templates") / template_name,
        ]
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return candidates[0]


def _sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^\w\s.-]+", "", value).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned or "employee"


def _cell_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _empty_uniform_issue_row() -> dict[str, str]:
    return {
        "description": "",
        "size": "",
        "quantity": "",
        "condition": "",
        "details": "",
        "returns": "",
        "cost": "",
    }


def _empty_uniform_issue_rows() -> list[dict[str, str]]:
    return [_empty_uniform_issue_row() for _ in range(UNIFORM_ISSUE_MAX_ROWS)]


def _uniform_issue_documents_query(db: Session, employee_id: int) -> list[EmployeeDocument]:
    return list(
        db.scalars(
            select(EmployeeDocument)
            .where(EmployeeDocument.employee_id == employee_id, EmployeeDocument.document_type == "uniform_issue")
            .order_by(EmployeeDocument.id.desc())
        ).all()
    )


def _latest_uniform_issue_document(db: Session, employee_id: int) -> EmployeeDocument | None:
    documents = _hidden_uniform_issue_documents_query(db, employee_id)
    return documents[0] if documents else None


def _uniform_issue_document_path(document: EmployeeDocument) -> Path:
    return _document_path(document.stored_filename)


def _uniform_issue_rows_from_workbook(workbook) -> list[dict[str, str]]:
    worksheet = workbook["Uniform"] if "Uniform" in workbook.sheetnames else workbook.active
    rows: list[dict[str, str]] = []

    for row_number in range(7, 7 + UNIFORM_ISSUE_MAX_ROWS):
        description = _cell_text(worksheet[f"A{row_number}"].value) or ""
        if not description and all(
            _cell_text(worksheet[f"{column}{row_number}"].value) is None
            for column in ["B", "C", "D", "E", "G", "H"]
        ):
            continue

        rows.append({
            "description": description,
            "size": _cell_text(worksheet[f"B{row_number}"].value) or "",
            "quantity": _cell_text(worksheet[f"C{row_number}"].value) or "",
            "condition": _cell_text(worksheet[f"D{row_number}"].value) or "",
            "details": _cell_text(worksheet[f"E{row_number}"].value) or "",
            "returns": _cell_text(worksheet[f"G{row_number}"].value) or "",
            "cost": _cell_text(worksheet[f"H{row_number}"].value) or "",
        })

    return rows


def _is_hidden_uniform_issue_workbook(document: EmployeeDocument) -> bool:
    return document.document_type == "uniform_issue" and (
        document.stored_filename.startswith("uniform_issue_") or document.original_filename.startswith("ISSUE-Uniforms - ")
    )


def _hidden_uniform_issue_documents_query(db: Session, employee_id: int) -> list[EmployeeDocument]:
    return [
        document
        for document in _uniform_issue_documents_query(db, employee_id)
        if _is_hidden_uniform_issue_workbook(document)
    ]


def _parse_iso_date(value: object) -> date | None:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _permanent_contract_start_date(employee: Employee) -> date | None:
    startup_data = employee.startup_data or {}
    contract_status = startup_data.get("Contract Status")
    if not isinstance(contract_status, dict):
        return None

    entries = contract_status.get("entries")
    candidate_dates: list[date] = []

    if isinstance(entries, list):
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            if entry.get("status") != "Permanent":
                continue
            parsed_date = _parse_iso_date(entry.get("date"))
            if parsed_date is not None:
                candidate_dates.append(parsed_date)

    if not candidate_dates and contract_status.get("status") == "Permanent":
        parsed_date = _parse_iso_date(contract_status.get("date"))
        if parsed_date is not None:
            candidate_dates.append(parsed_date)

    if not candidate_dates:
        return None

    return max(candidate_dates)


def _permit_expire_date(employee: Employee) -> date | None:
    startup_data = employee.startup_data or {}
    permit_status = startup_data.get("Permit Status")
    if not isinstance(permit_status, dict):
        return None

    return _parse_iso_date(permit_status.get("date"))


def calculate_annual_leave_balance(employee_id: int) -> tuple[float, float]:
    with SessionLocal() as db:
        employee = db.get(Employee, employee_id)
        if employee is None:
            return 0, 0

        permanent_start_date = _permanent_contract_start_date(employee)
        if permanent_start_date is None:
            return 0, 0

        today = date.today()
        months_employed = max(
            0,
            (today.year - permanent_start_date.year) * 12 + (today.month - permanent_start_date.month),
        )

        if employee.passport_id == "8601310127086":
            entitlement = months_employed * (20 / 12)
        else:
            entitlement = months_employed * 1.25

        used_days = (
            db.scalar(
                select(func.coalesce(func.sum(AnnualLeave.days_used), 0)).where(
                    AnnualLeave.employee_id == employee_id,
                    AnnualLeave.status == "Approved",
                )
            )
            or 0
        )
    balance = entitlement - float(used_days)
    return entitlement, balance


def calculate_sick_leave_balance(employee_id: int) -> tuple[float, float]:
    with SessionLocal() as db:
        employee = db.get(Employee, employee_id)
        if employee is None:
            return 0, 0

        permanent_start_date = _permanent_contract_start_date(employee)
        if permanent_start_date is None:
            return 0, 0

        today = date.today()
        days_employed = (today - permanent_start_date).days

        if days_employed < 180:
            entitlement = 6
            used_days = (
                db.scalar(
                    select(func.coalesce(func.sum(SickLeave.days_used), 0)).where(
                        SickLeave.employee_id == employee_id,
                        SickLeave.status == "Approved",
                    )
                )
                or 0
            )
            return entitlement, max(0, entitlement - float(used_days))

        days_after_six_months = days_employed - 180
        complete_cycles = days_after_six_months // 1095
        cycle_start_date = permanent_start_date + timedelta(days=180 + (complete_cycles * 1095))

        used_days = (
            db.scalar(
                select(func.coalesce(func.sum(SickLeave.days_used), 0)).where(
                    SickLeave.employee_id == employee_id,
                    SickLeave.status == "Approved",
                    SickLeave.start_date >= cycle_start_date,
                )
            )
            or 0
        )

        entitlement = 30
        if complete_cycles == 0:
            probation_used_days = (
                db.scalar(
                    select(func.coalesce(func.sum(SickLeave.days_used), 0)).where(
                        SickLeave.employee_id == employee_id,
                        SickLeave.status == "Approved",
                        SickLeave.start_date < (permanent_start_date + timedelta(days=180)),
                    )
                )
                or 0
            )
            balance = entitlement - float(probation_used_days) - float(used_days)
        else:
            balance = entitlement - float(used_days)

        return entitlement, max(0, balance)


def _annual_leave_response(leave: AnnualLeave, employee_name: str) -> AnnualLeaveResponse:
    return AnnualLeaveResponse(
        id=leave.id,
        employee_id=leave.employee_id,
        employee_name=employee_name,
        start_date=leave.start_date,
        end_date=leave.end_date,
        reason=leave.reason,
        days_used=float(leave.days_used),
        status=leave.status,
    )


def _sick_leave_response(leave: SickLeave, employee_name: str) -> SickLeaveResponse:
    return SickLeaveResponse(
        id=leave.id,
        employee_id=leave.employee_id,
        employee_name=employee_name,
        start_date=leave.start_date,
        end_date=leave.end_date,
        reason=leave.reason,
        days_used=float(leave.days_used),
        medical_cert=leave.medical_cert,
        status=leave.status,
    )


@app.get("/api/employees/{employee_id}/profile", response_model=EmployeeProfileResponse)
def get_employee_profile(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeProfileResponse:
    employee = _get_employee_or_404(db, employee_id)
    annual_entitlement, annual_balance = calculate_annual_leave_balance(employee_id)
    sick_entitlement, sick_balance = calculate_sick_leave_balance(employee_id)

    return EmployeeProfileResponse(
        id=employee.id,
        name=employee.name,
        department=employee.department,
        passport_id=employee.passport_id,
        archived=employee.archived,
        permanent_contract_start_date=_permanent_contract_start_date(employee),
        permit_expire_date=_permit_expire_date(employee),
        annual_leave_entitlement=annual_entitlement,
        annual_leave_balance=annual_balance,
        sick_leave_entitlement=sick_entitlement,
        sick_leave_balance=sick_balance,
    )


@app.get("/api/employees/{employee_id}/contract-files", response_model=list[ContractFileResponse])
def list_contract_files(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[ContractFileResponse]:
    _get_employee_or_404(db, employee_id)
    files = db.scalars(
        select(ContractFile).where(ContractFile.employee_id == employee_id).order_by(ContractFile.id.asc())
    ).all()
    return files


@app.post(
    "/api/employees/{employee_id}/contract-files",
    response_model=list[ContractFileResponse],
    status_code=status.HTTP_201_CREATED,
)
async def upload_contract_file(
    employee_id: int,
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[ContractFileResponse]:
    _get_employee_or_404(db, employee_id)

    for file in files:
        original_name = file.filename or ""
        suffix = Path(original_name).suffix.lower()
        if suffix not in ALLOWED_CONTRACT_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be a .pdf, .doc, .docx, .jpg, or .jpeg",
            )

    created_files: list[ContractFile] = []

    for file in files:
        original_name = file.filename or ""
        suffix = Path(original_name).suffix.lower()

        stored_filename = f"{uuid4().hex}{suffix}"
        destination = _contract_file_path(stored_filename)

        size_bytes = 0
        with destination.open("wb") as buffer:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                size_bytes += len(chunk)
                buffer.write(chunk)

        created_files.append(
            ContractFile(
                employee_id=employee_id,
                original_filename=original_name,
                stored_filename=stored_filename,
                content_type=file.content_type or "application/octet-stream",
                size_bytes=size_bytes,
            )
        )

    db.add_all(created_files)
    db.commit()
    for contract_file in created_files:
        db.refresh(contract_file)
    return created_files


@app.get("/api/contract-files/{file_id}/download")
def download_contract_file(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    contract_file = db.get(ContractFile, file_id)
    if contract_file is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    path = _contract_file_path(contract_file.stored_filename)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing from storage")

    return FileResponse(
        path,
        media_type=contract_file.content_type,
        filename=contract_file.original_filename,
    )


@app.get("/api/contract-files/{file_id}/view")
def view_contract_file(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    contract_file = db.get(ContractFile, file_id)
    if contract_file is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    path = _contract_file_path(contract_file.stored_filename)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing from storage")

    response = FileResponse(path, media_type=contract_file.content_type, filename=contract_file.original_filename)
    response.headers["Content-Disposition"] = f'inline; filename="{contract_file.original_filename}"'
    return response


@app.delete("/api/contract-files/{file_id}")
def delete_contract_file(
    file_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    contract_file = db.get(ContractFile, file_id)
    if contract_file is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    file_path = _contract_file_path(contract_file.stored_filename)
    db.delete(contract_file)
    db.commit()

    def remove_file(path: Path) -> None:
        if path.exists():
            path.unlink()

    background_tasks.add_task(remove_file, file_path)
    return {"message": "File deleted successfully"}


@app.get("/api/employees/{employee_id}/documents/{document_type}", response_model=list[EmployeeDocumentResponse])
def list_employee_documents(
    employee_id: int,
    document_type: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[EmployeeDocumentResponse]:
    _get_employee_or_404(db, employee_id)
    document_type = _normalize_document_type(document_type)
    documents = list(
        db.scalars(
        select(EmployeeDocument)
        .where(EmployeeDocument.employee_id == employee_id, EmployeeDocument.document_type == document_type)
        .order_by(EmployeeDocument.id.asc())
        ).all()
    )
    if document_type == "uniform_issue":
        documents = [document for document in documents if not _is_hidden_uniform_issue_workbook(document)]
    return documents


@app.post(
    "/api/employees/{employee_id}/documents/{document_type}",
    response_model=list[EmployeeDocumentResponse],
    status_code=status.HTTP_201_CREATED,
)
async def upload_employee_documents(
    employee_id: int,
    document_type: str,
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[EmployeeDocumentResponse]:
    _get_employee_or_404(db, employee_id)
    document_type = _normalize_document_type(document_type)

    for file in files:
        original_name = file.filename or ""
        suffix = Path(original_name).suffix.lower()
        if suffix not in ALLOWED_DOCUMENT_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be a .pdf, .doc, .docx, .jpg, .jpeg, .xls, or .xlsx",
            )

    created_documents: list[EmployeeDocument] = []

    for file in files:
        original_name = file.filename or ""
        suffix = Path(original_name).suffix.lower()

        stored_filename = f"{uuid4().hex}{suffix}"
        destination = _document_path(stored_filename)

        size_bytes = 0
        with destination.open("wb") as buffer:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                size_bytes += len(chunk)
                buffer.write(chunk)

        created_documents.append(
            EmployeeDocument(
                employee_id=employee_id,
                document_type=document_type,
                original_filename=original_name,
                stored_filename=stored_filename,
                content_type=file.content_type or "application/octet-stream",
                size_bytes=size_bytes,
            )
        )

    db.add_all(created_documents)
    db.commit()
    for document in created_documents:
        db.refresh(document)
    return created_documents


@app.post(
    "/api/employees/{employee_id}/documents/uniform_issue/save",
    response_model=EmployeeDocumentResponse,
    status_code=status.HTTP_201_CREATED,
)
def save_uniform_issue_document(
    employee_id: int,
    payload: UniformIssueSaveRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeDocumentResponse:
    employee = _get_employee_or_404(db, employee_id)
    cleaned_rows: list[dict[str, str]] = []
    for row in payload.rows:
        description = (row.description or "").strip()
        size = (row.size or "").strip()
        quantity = (row.quantity or "").strip()
        condition = (row.condition or "").strip()
        details = (row.details or "").strip()
        returns = (row.returns or "").strip()
        cost = (row.cost or "").strip()

        if not any([description, size, quantity, condition, details, returns, cost]):
            continue

        if not description:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Each row must have a description")
        if description not in UNIFORM_ISSUE_DESCRIPTION_OPTIONS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid uniform description")

        cleaned_rows.append(
            {
                "description": description,
                "size": size,
                "quantity": quantity,
                "condition": condition,
                "details": details,
                "returns": returns,
                "cost": cost,
            }
        )

    if len(cleaned_rows) > UNIFORM_ISSUE_MAX_ROWS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="You can only save up to 10 rows")

    template_path = _uniform_issue_template_path()
    if not template_path.exists():
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Uniform template is missing")

    workbook = load_workbook(template_path)
    worksheet = workbook["Uniform"] if "Uniform" in workbook.sheetnames else workbook.active

    worksheet["A4"] = employee.name

    for row_number in range(7, 7 + UNIFORM_ISSUE_MAX_ROWS):
        for column in ["A", "B", "C", "D", "E", "G", "H"]:
            worksheet[f"{column}{row_number}"] = None

    for index, row_data in enumerate(cleaned_rows):
        row_number = 7 + index
        worksheet[f"A{row_number}"] = row_data["description"]
        worksheet[f"B{row_number}"] = row_data["size"]
        worksheet[f"C{row_number}"] = row_data["quantity"]
        worksheet[f"D{row_number}"] = row_data["condition"]
        worksheet[f"E{row_number}"] = row_data["details"]
        worksheet[f"G{row_number}"] = row_data["returns"]
        worksheet[f"H{row_number}"] = row_data["cost"]

    safe_employee_name = _sanitize_filename_part(employee.name)
    original_filename = f"ISSUE-Uniforms - {safe_employee_name}.xlsx"
    stored_filename = f"uniform_issue_{employee.id}.xlsx"
    temp_output_path = UPLOAD_DIR / f".uniform_issue_{employee.id}_{uuid4().hex}.xlsx"
    output_path = UPLOAD_DIR / stored_filename

    workbook.save(temp_output_path)

    existing_documents = _hidden_uniform_issue_documents_query(db, employee_id)
    if existing_documents:
        existing_paths = [_uniform_issue_document_path(document) for document in existing_documents]
        for existing_document in existing_documents:
            db.delete(existing_document)
        db.commit()
        for existing_path in existing_paths:
            if existing_path.exists():
                existing_path.unlink()

    if output_path.exists():
        output_path.unlink()
    temp_output_path.replace(output_path)

    document = EmployeeDocument(
        employee_id=employee_id,
        document_type="uniform_issue",
        original_filename=original_filename,
        stored_filename=stored_filename,
        content_type=UNIFORM_ISSUE_MIME_TYPE,
        size_bytes=output_path.stat().st_size,
    )
    db.add(document)
    db.commit()
    db.refresh(document)
    return document


@app.get(
    "/api/employees/{employee_id}/uniform-issue-workbook",
    response_model=UniformIssueWorkbookResponse,
)
def get_uniform_issue_document(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> UniformIssueWorkbookResponse:
    employee = _get_employee_or_404(db, employee_id)
    document = _latest_uniform_issue_document(db, employee_id)
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uniform issue file not found")

    path = _uniform_issue_document_path(document)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing from storage")

    workbook = load_workbook(path)
    return UniformIssueWorkbookResponse(
        employee_name=employee.name,
        rows=_uniform_issue_rows_from_workbook(workbook),
    )


@app.get("/api/employees/{employee_id}/uniform-issue-workbook/download")
def download_uniform_issue_workbook(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    _get_employee_or_404(db, employee_id)
    document = _latest_uniform_issue_document(db, employee_id)
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uniform issue file not found")

    path = _uniform_issue_document_path(document)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing from storage")

    return FileResponse(path, media_type=document.content_type, filename=document.original_filename)


@app.get("/api/employees/{employee_id}/uniform-care-letter/download")
def download_uniform_care_letter(
    employee_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    employee = _get_employee_or_404(db, employee_id)
    template_path = _uniform_care_letter_template_path(employee.department)
    if not template_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uniform care letter template not found")

    return FileResponse(
        template_path,
        media_type=UNIFORM_CARE_LETTER_MIME_TYPE,
        filename=template_path.name,
    )


@app.get("/api/documents/{document_id}/download")
def download_employee_document(
    document_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    document = db.get(EmployeeDocument, document_id)
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    path = _document_path(document.stored_filename)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing from storage")

    return FileResponse(path, media_type=document.content_type, filename=document.original_filename)


@app.get("/api/documents/{document_id}/view")
def view_employee_document(
    document_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    document = db.get(EmployeeDocument, document_id)
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    path = _document_path(document.stored_filename)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing from storage")

    response = FileResponse(path, media_type=document.content_type, filename=document.original_filename)
    response.headers["Content-Disposition"] = f'inline; filename="{document.original_filename}"'
    return response


@app.delete("/api/documents/{document_id}")
def delete_employee_document(
    document_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    document = db.get(EmployeeDocument, document_id)
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    file_path = _document_path(document.stored_filename)
    db.delete(document)
    db.commit()

    def remove_file(path: Path) -> None:
        if path.exists():
            path.unlink()

    background_tasks.add_task(remove_file, file_path)
    return {"message": "File deleted successfully"}


def _medical_cert_path(filename: str) -> Path:
    return UPLOAD_DIR / filename


def _delete_medical_cert(filename: str | None) -> None:
    if not filename:
        return
    path = _medical_cert_path(filename)
    if path.exists():
        path.unlink()


@app.get("/api/annual-leave", response_model=list[AnnualLeaveResponse])
def list_annual_leave(
    employee_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[AnnualLeaveResponse]:
    query = (
        select(AnnualLeave, Employee.name)
        .join(Employee, AnnualLeave.employee_id == Employee.id)
        .where(Employee.archived.is_(False))
        .order_by(AnnualLeave.start_date.desc(), AnnualLeave.id.desc())
    )
    if employee_id is not None:
        query = query.where(AnnualLeave.employee_id == employee_id)

    rows = db.execute(query).all()
    return [_annual_leave_response(leave, employee_name) for leave, employee_name in rows]


@app.post("/api/annual-leave", response_model=AnnualLeaveResponse, status_code=status.HTTP_201_CREATED)
def create_annual_leave(
    payload: AnnualLeaveCreateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> AnnualLeaveResponse:
    employee = _get_employee_or_404(db, payload.employee_id)
    leave = AnnualLeave(
        employee_id=payload.employee_id,
        start_date=payload.start_date,
        end_date=payload.end_date,
        reason=payload.reason,
        days_used=payload.days_used,
        status=payload.status,
    )
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return _annual_leave_response(leave, employee.name)


@app.put("/api/annual-leave/{leave_id}", response_model=AnnualLeaveResponse)
def update_annual_leave(
    leave_id: int,
    payload: AnnualLeaveUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> AnnualLeaveResponse:
    leave = db.get(AnnualLeave, leave_id)
    if leave is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Annual leave not found")

    leave.start_date = payload.start_date
    leave.end_date = payload.end_date
    leave.reason = payload.reason
    leave.days_used = payload.days_used
    leave.status = payload.status
    db.add(leave)
    db.commit()
    db.refresh(leave)

    employee = _get_employee_or_404(db, leave.employee_id)
    return _annual_leave_response(leave, employee.name)


@app.delete("/api/annual-leave/{leave_id}")
def delete_annual_leave(
    leave_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    leave = db.get(AnnualLeave, leave_id)
    if leave is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Annual leave not found")

    db.delete(leave)
    db.commit()
    return {"message": "Annual leave deleted"}


async def _read_sick_leave_payload(request: Request) -> dict:
    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type:
        form = await request.form()
        try:
            start_date = date.fromisoformat(form.get("start_date"))
            end_date = date.fromisoformat(form.get("end_date"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Valid start and end dates are required")
        payload = {
            "employee_id": int(form.get("employee_id", 0)),
            "start_date": start_date,
            "end_date": end_date,
            "reason": form.get("reason", ""),
            "days_used": float(form.get("days_used", 0)),
            "status": form.get("status", "Approved"),
            "medical_cert": form.get("medical_cert", ""),
        }
        medical_cert_file = form.get("medical_cert_file")
        if medical_cert_file and getattr(medical_cert_file, "filename", ""):
            suffix = Path(medical_cert_file.filename).suffix.lower()
            if suffix not in ALLOWED_MEDICAL_CERT_EXTENSIONS:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Medical certificate must be a .pdf, .doc, .docx, .jpg, .jpeg, or .png",
                )
            stored_filename = f"{uuid4().hex}{suffix}"
            destination = _medical_cert_path(stored_filename)
            content = await medical_cert_file.read()
            destination.write_bytes(content)
            payload["medical_cert"] = stored_filename
        return payload

    data = await request.json()
    try:
        start_date = date.fromisoformat(data.get("start_date"))
        end_date = date.fromisoformat(data.get("end_date"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Valid start and end dates are required")
    return {
        "employee_id": int(data.get("employee_id", 0)),
        "start_date": start_date,
        "end_date": end_date,
        "reason": data.get("reason", ""),
        "days_used": float(data.get("days_used", 0)),
        "status": data.get("status", "Approved"),
        "medical_cert": data.get("medical_cert", ""),
    }


@app.get("/api/sick-leave", response_model=list[SickLeaveResponse])
def list_sick_leave(
    employee_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[SickLeaveResponse]:
    query = (
        select(SickLeave, Employee.name)
        .join(Employee, SickLeave.employee_id == Employee.id)
        .where(Employee.archived.is_(False))
        .order_by(SickLeave.start_date.desc(), SickLeave.id.desc())
    )
    if employee_id is not None:
        query = query.where(SickLeave.employee_id == employee_id)

    rows = db.execute(query).all()
    return [_sick_leave_response(leave, employee_name) for leave, employee_name in rows]


@app.post("/api/sick-leave", response_model=SickLeaveResponse, status_code=status.HTTP_201_CREATED)
async def create_sick_leave(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> SickLeaveResponse:
    payload = await _read_sick_leave_payload(request)
    employee = _get_employee_or_404(db, payload["employee_id"])
    leave = SickLeave(
        employee_id=payload["employee_id"],
        start_date=payload["start_date"],
        end_date=payload["end_date"],
        reason=payload["reason"],
        days_used=payload["days_used"],
        medical_cert=payload["medical_cert"] or None,
        status=payload["status"],
    )
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return _sick_leave_response(leave, employee.name)


@app.put("/api/sick-leave/{leave_id}", response_model=SickLeaveResponse)
async def update_sick_leave(
    leave_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> SickLeaveResponse:
    leave = db.get(SickLeave, leave_id)
    if leave is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sick leave not found")

    payload = await _read_sick_leave_payload(request)
    existing_medical_cert = leave.medical_cert
    new_medical_cert = payload["medical_cert"] or existing_medical_cert

    if payload["medical_cert"] == "":
        _delete_medical_cert(existing_medical_cert)
        new_medical_cert = None
    elif payload["medical_cert"] and payload["medical_cert"] != existing_medical_cert:
        _delete_medical_cert(existing_medical_cert)

    leave.start_date = payload["start_date"]
    leave.end_date = payload["end_date"]
    leave.reason = payload["reason"]
    leave.days_used = payload["days_used"]
    leave.medical_cert = new_medical_cert
    leave.status = payload["status"]
    db.add(leave)
    db.commit()
    db.refresh(leave)

    employee = _get_employee_or_404(db, leave.employee_id)
    return _sick_leave_response(leave, employee.name)


@app.delete("/api/sick-leave/{leave_id}")
def delete_sick_leave(
    leave_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    leave = db.get(SickLeave, leave_id)
    if leave is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sick leave not found")

    _delete_medical_cert(leave.medical_cert)
    db.delete(leave)
    db.commit()
    return {"message": "Sick leave deleted"}


@app.get("/api/sick-leave/{leave_id}/medical-cert/view")
def view_sick_leave_medical_cert(
    leave_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FileResponse:
    leave = db.get(SickLeave, leave_id)
    if leave is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sick leave not found")

    if not leave.medical_cert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medical certificate not found")

    path = _medical_cert_path(leave.medical_cert)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medical certificate missing from storage")

    return FileResponse(path, filename=Path(leave.medical_cert).name)


@app.get("/api/employee-leave/report", response_model=EmployeeLeaveReportResponse)
def employee_leave_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EmployeeLeaveReportResponse:
    employees = db.scalars(
        select(Employee)
        .where(Employee.archived.is_(False))
        .order_by(Employee.name.asc())
    ).all()
    annual = db.execute(
        select(AnnualLeave, Employee.name)
        .join(Employee, AnnualLeave.employee_id == Employee.id)
        .where(Employee.archived.is_(False))
        .order_by(AnnualLeave.start_date.desc(), AnnualLeave.id.desc())
    ).all()
    sick = db.execute(
        select(SickLeave, Employee.name)
        .join(Employee, SickLeave.employee_id == Employee.id)
        .where(Employee.archived.is_(False))
        .order_by(SickLeave.start_date.desc(), SickLeave.id.desc())
    ).all()

    summary = []
    for employee in employees:
        annual_entitlement, annual_balance = calculate_annual_leave_balance(employee.id)
        sick_entitlement, sick_balance = calculate_sick_leave_balance(employee.id)
        summary.append(
            EmployeeLeaveSummaryResponse(
                id=employee.id,
                name=employee.name,
                passport_id=employee.passport_id,
                permanent_contract_start_date=_permanent_contract_start_date(employee),
                annual_leave_entitlement=round(float(annual_entitlement), 2),
                annual_leave_balance=round(float(annual_balance), 2),
                sick_leave_entitlement=round(float(sick_entitlement), 2),
                sick_leave_balance=round(float(sick_balance), 2),
            )
        )

    return EmployeeLeaveReportResponse(
        employees=summary,
        annual=[_annual_leave_response(leave, employee_name) for leave, employee_name in annual],
        sick=[_sick_leave_response(leave, employee_name) for leave, employee_name in sick],
    )
